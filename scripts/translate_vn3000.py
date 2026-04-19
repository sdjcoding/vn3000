#!/usr/bin/env python3
"""Translate 3000 Vietnamese words from Excel into JP/EN + examples for all languages.

Requires: ANTHROPIC_API_KEY env var, openpyxl (auto-installed).
Progress-safe: resumes from data/vn3000_translations.json on restart.

Output:
  data/vn3000_translations.json  - raw translation results per item
  data/vn3000_cards.json         - final merged card data (Excel + translations)
"""
import os, sys, json, time, re, urllib.request, urllib.error
from pathlib import Path

API_KEY = os.environ.get("ANTHROPIC_API_KEY")
if not API_KEY:
    sys.exit("Set ANTHROPIC_API_KEY environment variable first:\n  export ANTHROPIC_API_KEY=sk-ant-...")

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"])
    import openpyxl

MODEL = "claude-sonnet-4-5"
BATCH_SIZE = 100

ROOT = Path(__file__).parent.parent
EXCEL = ROOT / "베트남어_3000단어.xlsx"
TRANSLATIONS = ROOT / "data" / "vn3000_translations.json"
CARDS = ROOT / "data" / "vn3000_cards.json"
TRANSLATIONS.parent.mkdir(exist_ok=True)

SYSTEM = """You translate Vietnamese vocabulary into Japanese and English, and generate short practical example sentences in four languages.

Input: JSON array of items, each {"no": int, "vn": str, "ko": str}.

Output: JSON array of objects, one per input, EXACTLY this shape:
{
  "no": <same no as input>,
  "vn_ex": "<Vietnamese example sentence, 5-12 words, natural & practical>",
  "ko_ex": "<Korean example sentence that matches vn_ex in meaning>",
  "jp": "<Japanese word translation. Wrap ALL kanji in <ruby>漢字<rt>かな</rt></ruby>>",
  "jp_ex": "<Japanese example sentence matching vn_ex. Wrap ALL kanji in <ruby>...<rt>...</rt></ruby>>",
  "en": "<English word translation>",
  "en_ex": "<English example sentence matching vn_ex>"
}

RULES:
- Return ONLY a valid JSON array. No markdown, no backticks, no prose.
- Preserve the exact "no" from input.
- If input "vn" is already a phrase/sentence, use it as-is for vn_ex.
- Keep examples short and natural. Prefer contexts a Korean businessperson working with Vietnamese partners would encounter.
- Japanese: EVERY kanji MUST have <ruby><rt>...</rt></ruby> furigana.
- Do not include the input fields in the output; only the output fields listed above.
"""

def call_api(items, max_attempts=3):
    payload = {
        "model": MODEL,
        "max_tokens": 16000,
        "system": SYSTEM,
        "messages": [{
            "role": "user",
            "content": json.dumps(items, ensure_ascii=False),
        }],
    }
    data = json.dumps(payload).encode("utf-8")
    for attempt in range(1, max_attempts + 1):
        try:
            req = urllib.request.Request(
                "https://api.anthropic.com/v1/messages",
                data=data,
                headers={
                    "content-type": "application/json",
                    "x-api-key": API_KEY,
                    "anthropic-version": "2023-06-01",
                },
            )
            with urllib.request.urlopen(req, timeout=300) as r:
                result = json.loads(r.read().decode("utf-8"))
            text = "".join(b["text"] for b in result["content"] if b["type"] == "text").strip()
            text = re.sub(r"^```\w*\n?", "", text).rstrip("` \n")
            parsed = json.loads(text)
            usage = result.get("usage", {})
            return parsed, usage
        except Exception as e:
            if attempt >= max_attempts:
                raise
            wait = 2 ** attempt
            print(f"    attempt {attempt} failed: {e} — retry in {wait}s", flush=True)
            time.sleep(wait)

def load_excel():
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        no, cat, lvl, vn, ko = (row + (None,) * 5)[:5]
        if no is None:
            continue
        rows.append({
            "no": int(no),
            "cat": (cat or "").strip(),
            "lvl": (lvl or "").strip(),
            "vn": (vn or "").strip(),
            "ko": (ko or "").strip(),
        })
    return rows

def main():
    excel_rows = load_excel()
    print(f"Loaded {len(excel_rows)} items from {EXCEL.name}")

    # Resume if output exists
    translations = {}
    if TRANSLATIONS.exists():
        for t in json.loads(TRANSLATIONS.read_text(encoding="utf-8")):
            translations[t["no"]] = t
        print(f"Resuming: {len(translations)} translations already saved")

    todo = [r for r in excel_rows if r["no"] not in translations]
    print(f"To translate: {len(todo)}")
    if not todo:
        print("All items translated. Skipping API calls.")
    else:
        total_batches = (len(todo) + BATCH_SIZE - 1) // BATCH_SIZE
        total_in = total_out = 0
        start = time.time()

        for i in range(0, len(todo), BATCH_SIZE):
            batch = todo[i:i + BATCH_SIZE]
            req_items = [{"no": r["no"], "vn": r["vn"], "ko": r["ko"]} for r in batch]
            batch_idx = i // BATCH_SIZE + 1
            t0 = time.time()
            try:
                results, usage = call_api(req_items)
            except Exception as e:
                print(f"  batch {batch_idx} FAILED after retries: {e}")
                print(f"  skipping no {batch[0]['no']}~{batch[-1]['no']}")
                continue

            # Validate & save
            got = 0
            for r in results:
                if isinstance(r, dict) and "no" in r:
                    translations[r["no"]] = r
                    got += 1
            out = [translations[k] for k in sorted(translations.keys())]
            TRANSLATIONS.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")

            total_in += usage.get("input_tokens", 0)
            total_out += usage.get("output_tokens", 0)
            elapsed = time.time() - t0
            print(f"  [{batch_idx}/{total_batches}] no {batch[0]['no']}~{batch[-1]['no']}: "
                  f"{got}/{len(batch)} items, {elapsed:.1f}s, "
                  f"in={usage.get('input_tokens','?')} out={usage.get('output_tokens','?')}",
                  flush=True)
            time.sleep(0.3)

        print(f"\nTotal tokens: in={total_in:,}, out={total_out:,}")
        est_cost = total_in * 3 / 1_000_000 + total_out * 15 / 1_000_000
        print(f"Estimated cost (Sonnet 4.5): ${est_cost:.2f}")
        print(f"Elapsed: {time.time() - start:.1f}s")

    # Build final card data (merge Excel + translations)
    card_data = []
    missing = 0
    for r in excel_rows:
        t = translations.get(r["no"], {})
        if not t:
            missing += 1
        card_data.append({
            "no": r["no"],
            "cat": r["cat"],
            "lvl": r["lvl"],
            "vn": r["vn"],
            "vn_ex": t.get("vn_ex", r["vn"]),
            "ko": r["ko"],
            "ko_ex": t.get("ko_ex", r["ko"]),
            "jp": t.get("jp", ""),
            "jp_ex": t.get("jp_ex", ""),
            "en": t.get("en", ""),
            "en_ex": t.get("en_ex", ""),
        })
    CARDS.write_text(json.dumps(card_data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nFinal cards: {len(card_data)} saved to {CARDS}")
    if missing:
        print(f"WARN: {missing} items missing translations (will show empty JP/EN fields)")

if __name__ == "__main__":
    main()
