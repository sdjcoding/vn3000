#!/usr/bin/env python3
"""Build data/vn3000_cards.json from 베트남어_3000단어.xlsx.

If data/vn3000_translations.json exists, merge its jp/en/examples in.
Otherwise, empty strings for those fields (UI shows VN + KO only).
"""
import json, sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"])
    import openpyxl

ROOT = Path(__file__).parent.parent
EXCEL = ROOT / "베트남어_3000단어.xlsx"
TRANSLATIONS = ROOT / "data" / "vn3000_translations.json"
CARDS = ROOT / "data" / "vn3000_cards.json"
CARDS.parent.mkdir(exist_ok=True)

# Load Excel
wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
excel_rows = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    no, cat, lvl, vn, ko = (row + (None,) * 5)[:5]
    if no is None:
        continue
    excel_rows.append({
        "no": int(no),
        "cat": (cat or "").strip(),
        "lvl": (lvl or "").strip(),
        "vn": (vn or "").strip(),
        "ko": (ko or "").strip(),
    })
print(f"Loaded {len(excel_rows)} rows from {EXCEL.name}")

# Load translations if present
translations = {}
if TRANSLATIONS.exists():
    for t in json.loads(TRANSLATIONS.read_text(encoding="utf-8")):
        translations[t["no"]] = t
    print(f"Merged {len(translations)} translations from {TRANSLATIONS.name}")
else:
    print(f"No translations yet ({TRANSLATIONS.name} missing) — JP/EN will be blank")

# Build card data
cards = []
for r in excel_rows:
    t = translations.get(r["no"], {})
    cards.append({
        "no": r["no"],
        "cat": r["cat"],
        "lvl": r["lvl"],
        "vn": r["vn"],
        "vn_ex": t.get("vn_ex", ""),
        "ko": r["ko"],
        "ko_ex": t.get("ko_ex", ""),
        "jp": t.get("jp", ""),
        "jp_ex": t.get("jp_ex", ""),
        "en": t.get("en", ""),
        "en_ex": t.get("en_ex", ""),
    })

CARDS.write_text(json.dumps(cards, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Saved {len(cards)} cards to {CARDS}")

# Print category summary
from collections import Counter
cat_counts = Counter(c["cat"] for c in cards)
print(f"\nCategories ({len(cat_counts)}):")
for cat, n in cat_counts.most_common():
    print(f"  {cat}: {n}")
